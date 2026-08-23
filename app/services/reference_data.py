"""Immutable free reference-data ingestion for R6 factor validation (#2912).

The source response is committed before normalization. A structural parser
failure therefore leaves the exact bytes and error in the database instead of
turning an upstream format change into an unreproducible job exception.
"""

from __future__ import annotations

import csv
import hashlib
import io
import zipfile
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Final, Literal

import httpx
import psycopg
from openpyxl import load_workbook
from psycopg.rows import dict_row

ReferenceSource = Literal["kenneth_french", "aqr", "fred"]
ReferenceUnit = Literal["decimal_return", "percent_per_annum", "binary_indicator"]

FRENCH_FIVE_FACTOR_URL: Final = (
    "https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Research_Data_5_Factors_2x3_CSV.zip"
)
FRENCH_MOMENTUM_URL: Final = "https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Momentum_Factor_CSV.zip"
AQR_VME_MONTHLY_URL: Final = (
    "https://www.aqr.com/-/media/AQR/Documents/Insights/Data-Sets/Value-and-Momentum-Everywhere-Factors-Monthly.xlsx"
)
FRED_CSV_URL: Final = "https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_key}"

FRENCH_PARSER_VERSION: Final = "kenneth-french-monthly-csv-v2"
AQR_PARSER_VERSION: Final = "aqr-vme-monthly-xlsx-v2"
FRED_PARSER_VERSION: Final = "fred-csv-v1"

_FRENCH_MISSING: Final = frozenset({Decimal("-99.99"), Decimal("-999")})
_AQR_HEADER: Final = (
    "DATE",
    "VAL",
    "MOM",
    "VAL^SS",
    "MOM^SS",
    "VAL^AA",
    "MOM^AA",
    "VALLS_VME_US90",
    "MOMLS_VME_US90",
    "VALLS_VME_UK90",
    "MOMLS_VME_UK90",
    "VALLS_VME_ROE90",
    "MOMLS_VME_ROE90",
    "VALLS_VME_JP90",
    "MOMLS_VME_JP90",
    "VALLS_VME_EQ",
    "MOMLS_VME_EQ",
    "VALLS_VME_FX",
    "MOMLS_VME_FX",
    "VALLS_VME_FI",
    "MOMLS_VME_FI",
    "VALLS_VME_COM",
    "MOMLS_VME_COM",
)


class ReferenceDataSourceError(ValueError):
    """A fetched response cannot satisfy its frozen source contract."""


@dataclass(frozen=True)
class ReferenceObservation:
    series_key: str
    observation_date: date
    value: Decimal
    unit: ReferenceUnit


@dataclass(frozen=True)
class ParsedReferenceData:
    observations: tuple[ReferenceObservation, ...]
    missing_count: int


Parser = Callable[[bytes], ParsedReferenceData]


@dataclass(frozen=True)
class ReferenceDatasetSpec:
    source: ReferenceSource
    dataset_key: str
    source_url: str
    parser_version: str
    parser: Parser


@dataclass(frozen=True)
class ReferenceRefreshReport:
    source: ReferenceSource
    dataset_key: str
    status: Literal["accepted", "not_modified", "unchanged"]
    snapshot_id: int
    response_sha256: str
    row_count: int
    missing_count: int
    first_observation: date | None
    last_observation: date | None


def _month_end(year: int, month: int) -> date:
    if month == 12:
        return date(year, 12, 31)
    return date(year, month + 1, 1).fromordinal(date(year, month + 1, 1).toordinal() - 1)


def _decimal(raw: object, *, context: str) -> Decimal:
    try:
        value = Decimal(str(raw).strip())
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ReferenceDataSourceError(f"{context}: value is not decimal") from exc
    if not value.is_finite():
        raise ReferenceDataSourceError(f"{context}: value must be finite")
    return value


def _validated(parsed: ParsedReferenceData) -> ParsedReferenceData:
    if not parsed.observations:
        raise ReferenceDataSourceError("source produced zero observations")
    seen: set[tuple[str, date]] = set()
    for item in parsed.observations:
        key = (item.series_key, item.observation_date)
        if key in seen:
            raise ReferenceDataSourceError(
                f"duplicate observation {item.series_key}/{item.observation_date.isoformat()}"
            )
        seen.add(key)
    return ParsedReferenceData(
        observations=tuple(sorted(parsed.observations, key=lambda item: (item.observation_date, item.series_key))),
        missing_count=parsed.missing_count,
    )


def parse_french_monthly_zip(
    payload: bytes,
    *,
    expected_series_keys: tuple[str, ...] | None = None,
) -> ParsedReferenceData:
    """Parse the one monthly table and stop before French's annual section."""
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            names = [name for name in archive.namelist() if not name.endswith("/")]
            if len(names) != 1:
                raise ReferenceDataSourceError(f"French ZIP contains {len(names)} files; expected one")
            text = archive.read(names[0]).decode("utf-8-sig")
    except (UnicodeDecodeError, zipfile.BadZipFile) as exc:
        raise ReferenceDataSourceError("French response is not a valid UTF-8 CSV ZIP") from exc

    rows = list(csv.reader(io.StringIO(text)))
    header_index = next(
        (index for index, row in enumerate(rows) if row and row[0].strip() == "" and len(row) > 1),
        None,
    )
    if header_index is None:
        raise ReferenceDataSourceError("French monthly header was not found")
    series_keys = tuple(cell.strip() for cell in rows[header_index][1:])
    if not series_keys or any(not key for key in series_keys) or len(set(series_keys)) != len(series_keys):
        raise ReferenceDataSourceError(f"French series header is invalid: {series_keys!r}")
    if expected_series_keys is not None and series_keys != expected_series_keys:
        raise ReferenceDataSourceError(f"French series header {series_keys!r}; expected {expected_series_keys!r}")

    observations: list[ReferenceObservation] = []
    missing_count = 0
    monthly_rows = 0
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not row:
            if monthly_rows:
                break
            continue
        stamp = row[0].strip()
        if len(stamp) != 6 or not stamp.isdigit():
            if monthly_rows:
                break
            continue
        if len(row) != len(series_keys) + 1:
            raise ReferenceDataSourceError(f"French row {row_number}: ragged monthly row")
        year = int(stamp[:4])
        month = int(stamp[4:])
        if not 1 <= month <= 12:
            raise ReferenceDataSourceError(f"French row {row_number}: invalid YYYYMM {stamp!r}")
        when = _month_end(year, month)
        monthly_rows += 1
        for series_key, raw in zip(series_keys, row[1:], strict=True):
            value = _decimal(raw, context=f"French row {row_number}/{series_key}")
            if value in _FRENCH_MISSING:
                missing_count += 1
                continue
            observations.append(ReferenceObservation(series_key, when, value / Decimal(100), "decimal_return"))
    return _validated(ParsedReferenceData(tuple(observations), missing_count))


def _aqr_date(raw: object, *, row_number: int) -> date:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        try:
            return datetime.strptime(raw.strip(), "%m/%d/%Y").date()
        except ValueError as exc:
            raise ReferenceDataSourceError(f"AQR row {row_number}: invalid DATE {raw!r}") from exc
    raise ReferenceDataSourceError(f"AQR row {row_number}: invalid DATE type {type(raw).__name__}")


def parse_aqr_vme_monthly(payload: bytes) -> ParsedReferenceData:
    """Parse AQR's named monthly sheet and exact 23-column factor surface."""
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:
        # openpyxl exposes several container/XML exception types. Convert them
        # at this external-data boundary while leaving later programmer errors
        # loud (iteration and validation occur outside this catch).
        raise ReferenceDataSourceError("AQR response is not a readable XLSX workbook") from exc
    try:
        if "VME Factors" not in workbook.sheetnames:
            raise ReferenceDataSourceError("AQR workbook has no 'VME Factors' worksheet")
        sheet = workbook["VME Factors"]
        header_row: int | None = None
        rows = sheet.iter_rows(values_only=True)
        for row_number, row in enumerate(rows, start=1):
            prefix = tuple(row[: len(_AQR_HEADER)])
            if prefix == _AQR_HEADER:
                if any(value is not None for value in row[len(_AQR_HEADER) :]):
                    raise ReferenceDataSourceError("AQR header has unexpected trailing columns")
                header_row = row_number
                break
            if row_number >= 100:
                break
        if header_row is None:
            raise ReferenceDataSourceError("AQR exact factor header was not found in the first 100 rows")

        observations: list[ReferenceObservation] = []
        missing_count = 0
        for row_number, row in enumerate(rows, start=header_row + 1):
            if not row or all(value is None or (isinstance(value, str) and not value.strip()) for value in row):
                continue
            if len(row) < len(_AQR_HEADER):
                raise ReferenceDataSourceError(f"AQR row {row_number}: ragged factor row")
            when = _aqr_date(row[0], row_number=row_number)
            for series_key, raw in zip(_AQR_HEADER[1:], row[1 : len(_AQR_HEADER)], strict=True):
                if raw is None or (isinstance(raw, str) and not raw.strip()):
                    missing_count += 1
                    continue
                value = _decimal(raw, context=f"AQR row {row_number}/{series_key}")
                observations.append(ReferenceObservation(series_key, when, value, "decimal_return"))
        return _validated(ParsedReferenceData(tuple(observations), missing_count))
    finally:
        workbook.close()


def parse_fred_csv(
    payload: bytes,
    *,
    series_key: str,
    unit: ReferenceUnit,
) -> ParsedReferenceData:
    """Parse one no-key FRED graph CSV with blank-as-missing semantics."""
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ReferenceDataSourceError("FRED response is not UTF-8 CSV") from exc
    reader = csv.DictReader(io.StringIO(text))
    expected = ("observation_date", series_key)
    if tuple(reader.fieldnames or ()) != expected:
        raise ReferenceDataSourceError(f"FRED header {reader.fieldnames!r}; expected {expected!r}")

    observations: list[ReferenceObservation] = []
    missing_count = 0
    for row_number, row in enumerate(reader, start=2):
        if None in row or any(row.get(key) is None for key in expected):
            raise ReferenceDataSourceError(f"FRED row {row_number}: ragged row")
        try:
            when = date.fromisoformat(str(row["observation_date"]).strip())
        except ValueError as exc:
            raise ReferenceDataSourceError(f"FRED row {row_number}: invalid observation_date") from exc
        raw = str(row[series_key]).strip()
        if not raw or raw == ".":
            missing_count += 1
            continue
        value = _decimal(raw, context=f"FRED row {row_number}/{series_key}")
        if unit == "binary_indicator" and value not in (Decimal(0), Decimal(1)):
            raise ReferenceDataSourceError(f"FRED row {row_number}/{series_key}: expected binary 0/1")
        observations.append(ReferenceObservation(series_key, when, value, unit))
    return _validated(ParsedReferenceData(tuple(observations), missing_count))


def _fred_parser(series_key: str, unit: ReferenceUnit) -> Parser:
    return lambda payload: parse_fred_csv(payload, series_key=series_key, unit=unit)


def _french_parser(expected_series_keys: tuple[str, ...]) -> Parser:
    return lambda payload: parse_french_monthly_zip(payload, expected_series_keys=expected_series_keys)


REFERENCE_DATASETS: Final[Mapping[str, ReferenceDatasetSpec]] = {
    "french_five_factor_monthly": ReferenceDatasetSpec(
        "kenneth_french",
        "french_five_factor_monthly",
        FRENCH_FIVE_FACTOR_URL,
        FRENCH_PARSER_VERSION,
        _french_parser(("Mkt-RF", "SMB", "HML", "RMW", "CMA", "RF")),
    ),
    "french_momentum_monthly": ReferenceDatasetSpec(
        "kenneth_french",
        "french_momentum_monthly",
        FRENCH_MOMENTUM_URL,
        FRENCH_PARSER_VERSION,
        _french_parser(("Mom",)),
    ),
    "aqr_vme_monthly": ReferenceDatasetSpec(
        "aqr",
        "aqr_vme_monthly",
        AQR_VME_MONTHLY_URL,
        AQR_PARSER_VERSION,
        parse_aqr_vme_monthly,
    ),
    "fred_dgs3mo": ReferenceDatasetSpec(
        "fred",
        "fred_dgs3mo",
        FRED_CSV_URL.format(series_key="DGS3MO"),
        FRED_PARSER_VERSION,
        _fred_parser("DGS3MO", "percent_per_annum"),
    ),
    "fred_usrec": ReferenceDatasetSpec(
        "fred",
        "fred_usrec",
        FRED_CSV_URL.format(series_key="USREC"),
        FRED_PARSER_VERSION,
        _fred_parser("USREC", "binary_indicator"),
    ),
}

FRENCH_DATASET_KEYS: Final = ("french_five_factor_monthly", "french_momentum_monthly")
AQR_DATASET_KEYS: Final = ("aqr_vme_monthly",)
FRED_DATASET_KEYS: Final = ("fred_dgs3mo", "fred_usrec")


def _latest_accepted(conn: psycopg.Connection[Any], spec: ReferenceDatasetSpec) -> Mapping[str, Any] | None:
    with conn.cursor(row_factory=dict_row) as cursor:
        cursor.execute(
            """
            SELECT snapshot_id, etag, last_modified, response_sha256,
                   row_count, missing_count, first_observation, last_observation
            FROM reference_data_snapshots
            WHERE source = %(source)s
              AND dataset_key = %(dataset_key)s
              AND parser_version = %(parser_version)s
              AND parse_status = 'accepted'
            ORDER BY fetched_at DESC, snapshot_id DESC
            LIMIT 1
            """,
            {
                "source": spec.source,
                "dataset_key": spec.dataset_key,
                "parser_version": spec.parser_version,
            },
        )
        return cursor.fetchone()


def _report_from_row(
    spec: ReferenceDatasetSpec,
    row: Mapping[str, Any],
    *,
    status: Literal["not_modified", "unchanged"],
) -> ReferenceRefreshReport:
    return ReferenceRefreshReport(
        source=spec.source,
        dataset_key=spec.dataset_key,
        status=status,
        snapshot_id=int(row["snapshot_id"]),
        response_sha256=str(row["response_sha256"]),
        row_count=int(row["row_count"]),
        missing_count=int(row["missing_count"]),
        first_observation=row["first_observation"],
        last_observation=row["last_observation"],
    )


def refresh_reference_dataset(
    conn: psycopg.Connection[Any],
    *,
    client: httpx.Client,
    spec: ReferenceDatasetSpec,
) -> ReferenceRefreshReport:
    """Fetch, raw-commit, parse and atomically accept one dataset snapshot."""
    if not conn.autocommit:
        raise RuntimeError("refresh_reference_dataset requires an autocommit connection")
    prior = _latest_accepted(conn, spec)
    headers: dict[str, str] = {}
    if prior is not None:
        if prior["etag"]:
            headers["If-None-Match"] = str(prior["etag"])
        if prior["last_modified"]:
            headers["If-Modified-Since"] = str(prior["last_modified"])
    response = client.get(spec.source_url, headers=headers)
    if response.status_code == 304:
        if prior is None:
            raise ReferenceDataSourceError("source returned 304 without an accepted prior snapshot")
        return _report_from_row(spec, prior, status="not_modified")
    response.raise_for_status()
    if not response.content:
        raise ReferenceDataSourceError("source returned an empty response body")

    response_sha256 = hashlib.sha256(response.content).hexdigest()
    with conn.transaction(), conn.cursor(row_factory=dict_row) as cursor:
        cursor.execute(
            """
            INSERT INTO reference_data_snapshots
                (source, dataset_key, source_url, etag, last_modified,
                 content_type, response_sha256, payload, parser_version)
            VALUES
                (%(source)s, %(dataset_key)s, %(source_url)s, %(etag)s,
                 %(last_modified)s, %(content_type)s, %(response_sha256)s,
                 %(payload)s, %(parser_version)s)
            ON CONFLICT (source, dataset_key, response_sha256, parser_version)
                DO NOTHING
            RETURNING snapshot_id, parse_status, parse_error, row_count,
                      missing_count, first_observation, last_observation
            """,
            {
                "source": spec.source,
                "dataset_key": spec.dataset_key,
                "source_url": spec.source_url,
                "etag": response.headers.get("ETag"),
                "last_modified": response.headers.get("Last-Modified"),
                "content_type": response.headers.get("Content-Type"),
                "response_sha256": response_sha256,
                "payload": response.content,
                "parser_version": spec.parser_version,
            },
        )
        snapshot = cursor.fetchone()
        if snapshot is None:
            cursor.execute(
                """
                SELECT snapshot_id, parse_status, parse_error, row_count,
                       missing_count, first_observation, last_observation
                FROM reference_data_snapshots
                WHERE source = %(source)s
                  AND dataset_key = %(dataset_key)s
                  AND response_sha256 = %(response_sha256)s
                  AND parser_version = %(parser_version)s
                ORDER BY snapshot_id DESC
                LIMIT 1
                """,
                {
                    "source": spec.source,
                    "dataset_key": spec.dataset_key,
                    "response_sha256": response_sha256,
                    "parser_version": spec.parser_version,
                },
            )
            snapshot = cursor.fetchone()
    if snapshot is None:
        raise RuntimeError("snapshot insert/conflict lookup returned no row")
    snapshot_id = int(snapshot["snapshot_id"])
    if snapshot["parse_status"] == "accepted":
        existing = {
            **snapshot,
            "response_sha256": response_sha256,
        }
        return _report_from_row(spec, existing, status="unchanged")
    if snapshot["parse_status"] == "rejected":
        raise ReferenceDataSourceError(
            f"snapshot {snapshot_id} was already rejected by {spec.parser_version}: {snapshot['parse_error']}"
        )

    try:
        parsed = spec.parser(response.content)
    except ReferenceDataSourceError as exc:
        with conn.transaction():
            result = conn.execute(
                """
                UPDATE reference_data_snapshots
                SET parse_status = 'rejected', parse_error = %(error)s,
                    parsed_at = now()
                WHERE snapshot_id = %(snapshot_id)s AND parse_status = 'pending'
                """,
                {"snapshot_id": snapshot_id, "error": str(exc)[:4000]},
            )
            if result.rowcount != 1:
                raise RuntimeError(f"expected one pending snapshot {snapshot_id} to reject") from exc
        raise

    dates = [item.observation_date for item in parsed.observations]
    with conn.transaction():
        with conn.cursor() as cursor:
            cursor.executemany(
                """
                INSERT INTO reference_data_observations
                    (snapshot_id, series_key, observation_date, value, unit)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (snapshot_id, series_key, observation_date) DO NOTHING
                """,
                [
                    (snapshot_id, item.series_key, item.observation_date, item.value, item.unit)
                    for item in parsed.observations
                ],
            )
        result = conn.execute(
            """
            UPDATE reference_data_snapshots
            SET parse_status = 'accepted', parse_error = NULL, parsed_at = now(),
                row_count = %(row_count)s, missing_count = %(missing_count)s,
                first_observation = %(first)s, last_observation = %(last)s
            WHERE snapshot_id = %(snapshot_id)s AND parse_status = 'pending'
            """,
            {
                "snapshot_id": snapshot_id,
                "row_count": len(parsed.observations),
                "missing_count": parsed.missing_count,
                "first": min(dates),
                "last": max(dates),
            },
        )
        if result.rowcount != 1:
            raise RuntimeError(f"expected one pending snapshot {snapshot_id} to accept")
    return ReferenceRefreshReport(
        source=spec.source,
        dataset_key=spec.dataset_key,
        status="accepted",
        snapshot_id=snapshot_id,
        response_sha256=response_sha256,
        row_count=len(parsed.observations),
        missing_count=parsed.missing_count,
        first_observation=min(dates),
        last_observation=max(dates),
    )


def refresh_reference_group(
    conn: psycopg.Connection[Any],
    *,
    client: httpx.Client,
    dataset_keys: Sequence[str],
) -> tuple[ReferenceRefreshReport, ...]:
    """Refresh every member of one source group, then surface any failures."""
    unknown = [key for key in dataset_keys if key not in REFERENCE_DATASETS]
    if unknown:
        raise ValueError(f"unknown reference dataset key(s): {unknown}")
    reports: list[ReferenceRefreshReport] = []
    failures: list[Exception] = []
    for key in dataset_keys:
        try:
            reports.append(refresh_reference_dataset(conn, client=client, spec=REFERENCE_DATASETS[key]))
        except Exception as exc:
            exc.add_note(f"reference dataset: {key}")
            failures.append(exc)
    if failures:
        keys = ", ".join(dataset_keys)
        raise ExceptionGroup(f"one or more reference refreshes failed in group [{keys}]", failures)
    return tuple(reports)


__all__ = [
    "AQR_DATASET_KEYS",
    "AQR_PARSER_VERSION",
    "AQR_VME_MONTHLY_URL",
    "FRED_DATASET_KEYS",
    "FRED_PARSER_VERSION",
    "FRENCH_DATASET_KEYS",
    "FRENCH_PARSER_VERSION",
    "REFERENCE_DATASETS",
    "ParsedReferenceData",
    "ReferenceDataSourceError",
    "ReferenceDatasetSpec",
    "ReferenceObservation",
    "ReferenceRefreshReport",
    "parse_aqr_vme_monthly",
    "parse_fred_csv",
    "parse_french_monthly_zip",
    "refresh_reference_dataset",
    "refresh_reference_group",
]
